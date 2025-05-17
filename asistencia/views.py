import base64
import io
import qrcode

from datetime import datetime, date, timedelta
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response

from PIL import Image
from django.core.files.base import ContentFile
import openpyxl

from .models import (
    Proyecto, Trabajador, Asistencia,
    Dispositivo, SesionAsistencia
)
from .serializers import ProyectoSerializer, TrabajadorSerializer, AsistenciaSerializer


# =======================================================
# ViewSets para API REST
# =======================================================
class ProyectoViewSet(viewsets.ModelViewSet):
    queryset = Proyecto.objects.all()
    serializer_class = ProyectoSerializer


class TrabajadorViewSet(viewsets.ModelViewSet):
    queryset = Trabajador.objects.all()
    serializer_class = TrabajadorSerializer


class AsistenciaViewSet(viewsets.ModelViewSet):
    queryset = Asistencia.objects.all()
    serializer_class = AsistenciaSerializer


# =======================================================
# Vistas web y API complementarias
# =======================================================
def asistencia_elegir_proyecto_view(request):
    proyectos = Proyecto.objects.all().order_by('nombre')
    project_id = request.GET.get('project_id')
    selected_project = None
    trabajadores = []

    if project_id:
        selected_project = get_object_or_404(Proyecto, pk=project_id)
        trabajadores = selected_project.trabajadores.all().order_by('apellido_paterno', 'apellido_materno')

    return render(request, 'asistencia/asistencia_elegir.html', {
        'proyectos': proyectos,
        'selected_project': selected_project,
        'trabajadores': trabajadores,
    })


class RegistrarAsistenciaView(APIView):
    """
    Endpoint para registrar asistencia diaria vía JSON:
    { project, date, asistencias: [ {trabajador, presente}, … ] }
    """
    def post(self, request):
        data = request.data
        project_id     = data.get("project")
        fecha_str      = data.get("date")
        asistencias_data = data.get("asistencias", [])

        if not project_id or not fecha_str:
            return Response({"error": "Project and date are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Date format must be YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        proyecto = get_object_or_404(Proyecto, pk=project_id)
        results = []

        for item in asistencias_data:
            trab_id = item.get("trabajador")
            presente = item.get("presente", False)
            try:
                trab = Trabajador.objects.get(pk=trab_id)
            except Trabajador.DoesNotExist:
                results.append({"trabajador": trab_id, "error": "Trabajador not found."})
                continue

            asistencia, created = Asistencia.objects.update_or_create(
                trabajador=trab,
                proyecto=proyecto,
                fecha=fecha,
                defaults={"presente": presente}
            )
            results.append({"trabajador": trab_id, "created": created, "presente": presente})

        return Response({"status": "success", "results": results}, status=status.HTTP_200_OK)


class ExportarAsistenciaExcelView(APIView):
    """
    Versión simple: exporta asistencia a Excel con project_id, start_date, end_date.
    """
    def get(self, request):
        project_id     = request.GET.get('project_id')
        start_date_str = request.GET.get('start_date')
        end_date_str   = request.GET.get('end_date')

        if not project_id or not start_date_str or not end_date_str:
            return HttpResponse("project_id, start_date y end_date son requeridos", status=400)

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date   = datetime.strptime(end_date_str,   "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Formato de fecha debe ser YYYY-MM-DD.", status=400)

        proyecto     = get_object_or_404(Proyecto, pk=project_id)
        trabajadores = proyecto.trabajadores.all().order_by('apellido_paterno', 'apellido_materno')
        delta        = end_date - start_date
        fechas       = [start_date + timedelta(days=i) for i in range(delta.days + 1)]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asistencia"

        # Encabezados
        ws.cell(row=1, column=1, value="Nombre Completo")
        ws.cell(row=1, column=2, value="Categoría")
        for idx, f in enumerate(fechas, start=3):
            dia = {0:"L",1:"M",2:"MX",3:"J",4:"V",5:"S",6:"D"}[f.weekday()]
            ws.cell(row=1, column=idx, value=f"{dia} {f.strftime('%d/%m/%y')}")

        # Datos
        for r, trab in enumerate(trabajadores, start=2):
            nombre = f"{trab.nombre} {trab.apellido_paterno} {trab.apellido_materno}"
            ws.cell(row=r, column=1, value=nombre)
            ws.cell(row=r, column=2, value=trab.categoria)
            for c, f in enumerate(fechas, start=3):
                try:
                    asi = Asistencia.objects.get(trabajador=trab, proyecto=proyecto, fecha=f)
                    ws.cell(row=r, column=c, value="✓" if asi.presente else "X")
                except Asistencia.DoesNotExist:
                    ws.cell(row=r, column=c, value="")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        fname = f"asistencia_{project_id}_{start_date_str}_to_{end_date_str}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


@login_required
def asistencia_view(request, project_id):
    proyecto     = get_object_or_404(Proyecto, pk=project_id)
    trabajadores = proyecto.trabajadores.all().order_by('apellido_paterno', 'apellido_materno')
    fecha_actual = date.today().strftime("%Y-%m-%d")
    return render(request, 'asistencia/asistencia.html', {
        'proyecto': proyecto,
        'trabajadores': trabajadores,
        'fecha': fecha_actual,
    })


def registrar_asistencia_form_view(request):
    if request.method == "POST":
        project_id = request.POST.get("project_id")
        fecha_str  = request.POST.get("fecha")
        if not project_id or not fecha_str:
            messages.error(request, "Faltan datos para registrar la asistencia.")
            return redirect('alta_trabajador')

        try:
            fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('alta_trabajador')

        proyecto     = get_object_or_404(Proyecto, pk=project_id)
        trabajadores = proyecto.trabajadores.all()
        for trab in trabajadores:
            if f"asistencia_{trab.id}" in request.POST:
                Asistencia.objects.update_or_create(
                    trabajador=trab,
                    proyecto=proyecto,
                    fecha=fecha,
                    defaults={"presente": True}
                )
        messages.success(request, "Asistencia registrada con éxito.")
        return redirect(f"{redirect('asistencia-elegir')}?project_id={project_id}")
    return redirect('alta_trabajador')


def alta_trabajador_view(request):
    proyectos = Proyecto.objects.all()
    if request.method == "POST":
        nombre           = request.POST.get('nombre')
        apellido_paterno = request.POST.get('apellido_paterno')
        apellido_materno = request.POST.get('apellido_materno')
        categoria        = request.POST.get('categoria')
        telefono         = request.POST.get('telefono')
        curp             = request.POST.get('curp')
        nss              = request.POST.get('nss')
        proyectos_ids    = request.POST.getlist('proyectos')

        trabajador = Trabajador(
            nombre=nombre,
            apellido_paterno=apellido_paterno,
            apellido_materno=apellido_materno,
            categoria=categoria,
            telefono=telefono,
            curp=curp,
            nss=nss
        )

        # Procesar fotografía en base64
        foto_data = request.POST.get('fotografia')
        if foto_data:
            try:
                fmt, imgstr = foto_data.split(';base64,')
                img_data     = base64.b64decode(imgstr)
                image        = Image.open(io.BytesIO(img_data))
                image        = image.resize((300, 400))
                buf          = io.BytesIO()
                image.save(buf, format='PNG')
                trabajador.fotografia.save(
                    "trabajador_foto.png",
                    ContentFile(buf.getvalue()),
                    save=False
                )
            except Exception:
                messages.error(request, "Error al procesar la fotografía.")

        trabajador.save()
        if proyectos_ids:
            trabajador.proyectos.set(proyectos_ids)

        messages.success(request, "Trabajador dado de alta correctamente.")

        # Generar QR para la credencial
        qr_data     = f"{trabajador.id}-{trabajador.nombre} {trabajador.apellido_paterno}"
        qr_img      = qrcode.make(qr_data)
        buffer_qr   = io.BytesIO()
        qr_img.save(buffer_qr, format='PNG')
        qr_b64      = base64.b64encode(buffer_qr.getvalue()).decode("utf-8")

        return render(request, 'asistencia/alta_trabajador.html', {
            'proyectos': proyectos,
            'nueva_credencial': trabajador,
            'qr_image_base64': qr_b64,
        })

    return render(request, 'asistencia/alta_trabajador.html', {'proyectos': proyectos})


class RegistrarAsistenciaQRView(APIView):
    """
    Registra asistencia vía QR y device_id:
      - Solo dispositivos autorizados.
      - Primer escaneo del día fija hora_base.
      - Permite hasta 60 min.
      - ≤10min: puntual; 11–40: retardo_leve; 41–60: retardo_alto; >60: rechazado.
    """
    def get(self, request, trabajador_id):
        device_id = request.GET.get('device_id')
        if not device_id:
            return Response({'error': 'device_id es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        # 1) validar dispositivo
        try:
            disp = Dispositivo.objects.get(device_id=device_id)
        except Dispositivo.DoesNotExist:
            return Response({'error': 'Dispositivo no autorizado.'}, status=status.HTTP_403_FORBIDDEN)

        # 2) obtener trabajador + proyecto
        trab = get_object_or_404(Trabajador, pk=trabajador_id)
        proj = trab.proyectos.first()
        if proj not in disp.proyectos.all():
            return Response({'error': 'Device no autorizado para este proyecto.'}, status=status.HTTP_403_FORBIDDEN)

        hoy   = date.today()
        ahora = timezone.now()

        # 3) sesión diaria
        sesion, creada = SesionAsistencia.objects.get_or_create(
            dispositivo=disp, proyecto=proj, fecha=hoy,
            defaults={'hora_base': ahora}
        )
        minutos = 0 if creada else (ahora - sesion.hora_base).total_seconds() / 60

        # 4) clasificar o rechazar
        if minutos > 60:
            return Response({'error': 'Tiempo excedido (>60 min).'}, status=status.HTTP_400_BAD_REQUEST)
        elif minutos <= 10:
            tipo = 'puntual'
        elif minutos <= 40:
            tipo = 'retardo_leve'
        else:
            tipo = 'retardo_alto'

        # 5) grabar asistencia
        Asistencia.objects.update_or_create(
            trabajador=trab, proyecto=proj, fecha=hoy,
            defaults={'presente': True, 'tipo_retraso': tipo}
        )
        return Response({'message': 'Asistencia registrada.', 'tipo_retraso': tipo}, status=status.HTTP_200_OK)


def bienvenido_view(request):
    return render(request, 'bienvenido.html')
from django.shortcuts import render

def scan_offline_view(request):
    """
    Página offline que abre la cámara y escanea el QR.
    """
    return render(request, "asistencia/scan_offline.html")
